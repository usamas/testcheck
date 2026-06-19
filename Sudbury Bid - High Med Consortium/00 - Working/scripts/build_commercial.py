#!/usr/bin/env python3
"""Populate the known/structural parts of Document 5D (Commercial Schedule) and
Document 5C (FRT), and flag financial cells that require confirmed figures.

We deliberately DO NOT fabricate salary, locum, global-sum or audited-accounts
figures. Cost cells are marked 'TBC' against the relevant clarification (CQ), so
the finance lead completes them from confirmed data before submission.
"""
import os
import shutil
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, "..", "..", ".."))
SRC = os.path.join(ROOT, "Sudbury surgery")
TRD = os.path.normpath(os.path.join(HERE, "..", "..", "01 - Tender Response Documents"))

D_TPL = os.path.join(SRC, "Document 5 D - Commercial Schedule for Lot 4 Sudbury Surgery.xlsx")
C_TPL = os.path.join(SRC, "Document 5 C - Financial Ratio Template (FRT) for Lot 4 Sudbury Surgery.xlsx")
D_OUT = os.path.join(TRD, "Document 5D - Commercial Schedule.xlsx")
C_OUT = os.path.join(TRD, "Document 5C - Financial Ratio Template (FRT).xlsx")

# 17 positions in scope (Document 4, Lot 4) + PCN ARRS line
WORKFORCE = [
    "Lead GP (transferring)",
    "Salaried GP (37.5 hrs) - vacancy to recruit",
    "Salaried GP (20 hrs) - vacancy to recruit",
    "Practice Nurse (37.5 hrs) - vacancy to recruit",
    "Healthcare Assistant (transferring)",
    "Phlebotomist 1 (transferring)",
    "Phlebotomist 2 (transferring)",
    "Practice Manager (transferring)",
    "Deputy Practice Manager (transferring)",
    "Reception/Admin 1 (transferring)",
    "Reception/Admin 2 (transferring)",
    "Reception/Admin 3 (transferring)",
    "Reception/Admin 4 (transferring)",
    "Reception/Admin 5 (transferring)",
    "Reception/Admin 6 (transferring)",
    "Reception/Admin 7 (transferring)",
    "PCN ARRS (pharmacist/FCP/SPLW/HWBC) - network-allocated",
]

CQ_NOTE = ("DRAFT - cost figures TBC pending clarifications: global sum/affordability "
           "(CQ5/CQ22), LES/DES revenue (CQ52), locum costs (CQ53), estates/service "
           "charges & NIA (CQ9/CQ21/CQ44/CQ48), ARRS WTE (CQ50), QOF (CQ51). "
           "Verified estates: rent ~GBP 337,824/yr; rates ~GBP 12,906/yr. "
           "Annual Guidance Value GBP 1,640,000. Complete from confirmed data before submission.")


def populate_5d():
    shutil.copyfile(D_TPL, D_OUT)
    wb = openpyxl.load_workbook(D_OUT)
    ws = wb["Commercial Schedule"]
    # Workforce roster names (section 2: rows 12..28 in column B)
    start = 12
    for i, role in enumerate(WORKFORCE):
        ws.cell(start + i, 2).value = role
    # Patient activity (7,788 raw list) across years 1-5 if the activity cell is open
    # Row 8 is authority-defined; we add a bidder note rather than overwrite it.
    # Prominent draft note in an unused cell.
    ws.cell(2, 1).value = "BIDDER: High Med Ltd (lead) + Dr Singh Hammond Road - Lot 4 Sudbury Surgery (E84685)"
    note_row = 145
    ws.cell(note_row, 1).value = "BIDDER NOTE:"
    ws.cell(note_row, 2).value = CQ_NOTE
    ws.cell(note_row + 1, 1).value = "List size:"
    ws.cell(note_row + 1, 2).value = "7,788 raw registered (1 Apr 2026); recalibrate to Carr-Hill weighted at award (CQ47)."
    wb.save(D_OUT)
    print(f"5D populated: {len(WORKFORCE)} workforce roles + bidder/CQ notes")


def populate_5c():
    shutil.copyfile(C_TPL, C_OUT)
    wb = openpyxl.load_workbook(C_OUT)
    ws = wb["For BIDDER COMPLETION"]
    # Two completion blocks: first for one bidder member, second for the other.
    ws["B2"] = "High Med Ltd (Lead Bidder)"
    ws["C5"] = "Enter FY from audited accounts (High Med Ltd) - see 03 Supporting Evidence"
    ws["C7"] = "TBC - from audited accounts (do not estimate)"
    ws["C16"] = "Share of contract value % - to confirm at consortium agreement"
    # Second block (row 31 onwards) for the key party
    ws["A30"] = "Bidder Member 2: Dr Singh Hammond Road (key party)"
    ws["C31"] = "Enter FY from audited accounts (Dr Singh Hammond Road) - see 03 Supporting Evidence"
    # Guidance note (A54 is outside the merged ranges)
    ws["A54"] = ("FRT to be completed per bidder member from audited accounts "
                 "(31.3.23 / 31.3.24 in Supporting Evidence). Figures must not be estimated. "
                 "Annual Guidance Value GBP 1,640,000.")
    wb.save(C_OUT)
    print("5C populated: bidder names + completion guidance (financial figures left for confirmed entry)")


if __name__ == "__main__":
    populate_5d()
    populate_5c()
    print("Commercial schedules written to 01 - Tender Response Documents/")
